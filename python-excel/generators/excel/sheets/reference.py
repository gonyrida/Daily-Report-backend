from ...excel.templates import copy_row, copy_merged_cells
from ...common.helpers import write_to_merged_safe
from openpyxl.styles import Alignment
from ...common.config import (
    ENTRY_BLOCK_START, 
    ENTRY_BLOCK_END, 
    ENTRY_HEIGHT, 
    DATA_COLUMNS, 
    FOOTER_COLUMNS
)
from ..images import process_and_insert_images
from openpyxl.worksheet.pagebreak import Break

def prepare_entry_block(ws, current_row, template_start, template_end):
    """
    Clones the template structure into the new location and 
    forces the row height specifically for image fitting.
    """
    # Only clone if we aren't currently sitting on the template row itself
    if current_row != template_start:
        offset = current_row - template_start
        for r in range(template_start, template_end + 1):
            copy_row(ws, r, current_row + (r - template_start))
        copy_merged_cells(ws, template_start, template_end, offset)
    
    # Force the Image Row (usually current_row + 1) to be large enough for photos
    # 3.7 inches * 72 points/inch = ~267 points
    ws.row_dimensions[current_row + 1].height = 267

def write_footers(ws, row, entry, footer_columns):
    """
    Writes footer data (like labels or dates) into the 
    pre-merged footer cells in the Excel sheet.
    """
    footers = entry.get("footers", ["", ""])
    for idx, text in enumerate(footers):
        if idx < len(footer_columns):
            # Using the common helper to ensure we don't break merged cells
            write_to_merged_safe(ws, row, footer_columns[idx], text)

def fill_reference_sheet(ws, reference_entries, table_title="PHOTO REFERENCE"):

    # """Fill the reference sheet with reference sections"""
    
    # print(f"DEBUG: fill_reference_sheet received: {reference_entries}")
    # print(f"DEBUG: Number of sections: {len(reference_entries) if reference_entries else 0}")
    
    # if not reference_entries:
    #     print("DEBUG: No reference data provided")
    #     return
    
    # # Check if reference_entries is a list
    # if isinstance(reference_entries, list):
    #     print("DEBUG: reference_entries is a list")
    #     sections = reference_entries
    # else:
    #     print(f"DEBUG: reference_entries is not a list, type: {type(reference_entries)}")
    #     return
    
    # # Process each section
    # current_row = 1
    # for section in sections:
    #     print(f"DEBUG: Processing section: {section}")
        
    #     # Write section title
    #     ws[f'A{current_row}'] = section.get('section_title', '')
    #     print(f"DEBUG: Wrote section title to A{current_row}: {section.get('section_title', '')}")
        
    #     # Process images
    #     images = section.get('images', [])
    #     if images:
    #         for i, image_path in enumerate(images, 1):
    #             ws[f'B{current_row}'] = f"Image {i}"
    #             ws[f'C{current_row}'] = image_path
    #             print(f"DEBUG: Wrote image {i} to C{current_row}: {image_path}")
        
    #     # Process footers
    #     footers = section.get('footers', [])
    #     for i, footer in enumerate(footers, 1):
    #             ws[f'E{current_row}'] = footer
    #             print(f"DEBUG: Wrote footer {i} to E{current_row}: {footer}")
        
    #     current_row += 3  # Add spacing between sections
        
    # print(f"DEBUG: Final current_row: {current_row}")

    # # 1. Header setup
    write_to_merged_safe(ws, 3, 2, f"◙ {table_title}")
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')

    current_row = ENTRY_BLOCK_START # Row 6
    last_section = None
    image_cache = {}
    entries_on_current_page = 0 

    for entry in reference_entries:
        current_section = entry.get("section_title")

        # --- CHANGE 1: SECTION HEADER LOGIC ---
        if current_section and current_section != last_section:
            # If we reached 4 entries, break BEFORE printing the next Section Title
            if entries_on_current_page >= 4 and last_section is not None:
                ws.row_breaks.append(Break(id=current_row - 1))
                entries_on_current_page = 0 

            if last_section is not None:
                copy_row(ws, 5, current_row)
                copy_merged_cells(ws, 5, 5, current_row - 5)
                write_to_merged_safe(ws, current_row, 2, current_section)
                current_row += 1 
            else:
                # First section of the whole document
                write_to_merged_safe(ws, 5, 2, current_section)
            
            last_section = current_section

        # --- CHANGE 2: ENTRY BREAK LOGIC ---
        # If an entry (not a section title) is about to be the 5th item, break first
        if entries_on_current_page >= 4:
            ws.row_breaks.append(Break(id=current_row - 1))
            entries_on_current_page = 0

        # 4. IMAGE BLOCK LOGIC (Remains the same)
        prepare_entry_block(ws, current_row, ENTRY_BLOCK_START, ENTRY_BLOCK_END)
        process_and_insert_images(ws, entry, image_cache, current_row + 1, DATA_COLUMNS)
        write_footers(ws, current_row + 3, entry, FOOTER_COLUMNS)
        
        current_row += ENTRY_HEIGHT
        entries_on_current_page += 1 

    # Final Setup
    ws.print_title_rows = '1:4'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0