from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from ...common.helpers import write_wrapped_rows, to_num
from copy import copy
from ..templates import copy_cell_style

def fill_report_header(ws, data):
    # We are switching from "B" to "A" because "B" is a read-only MergedCell
    
    # Project Name (B7)
    ws["B7"].value = f"Project Name : {data.get('projectName', '')}"

    # Weather (B8)
    weather_am = data.get('weatherAM', "")
    weather_pm = data.get('weatherPM', "")
    ws["B8"].value = f"Weather          : AM {weather_am}  |  PM {weather_pm}"

    # Temperature (B9)
    temp_am = f"{data.get('tempAM')}°C" if data.get('tempAM') else ""
    temp_pm = f"{data.get('tempPM')}°C" if data.get('tempPM') else ""
    ws["B9"].value = f"Temperature  : AM {temp_am}    |  PM {temp_pm}"

    # Date (I9) - Usually not merged with H, so I9 should be fine
    report_date = data.get('reportDate')
    if report_date:
        try:
            dt_str = str(report_date).replace('Z', '')
            dt = datetime.fromisoformat(dt_str)
            cell = ws["I9"]
            cell.value = dt
            
            # Apply the exact format from your template
            # The [$-F800] ensures it follows the system's "Long Date" style
            cell.number_format = '[$-F800]dddd, mmmm dd, yyyy'
            
            # Since "Wednesday, September 25, 2026" is very long, 
            # make sure it aligns to the left so it doesn't hide behind other cells
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
        except Exception as e:
            print(f"Date Error: {e}")
            ws["I9"].value = report_date

def fill_activities(ws, data):
    report_style = Alignment(horizontal='left', vertical='top', wrap_text=True)
    normal_font = Font(bold=False, name='Arial', size=10)

    # Activities Today (B12 to B21)
    write_wrapped_rows(ws, start_row=12, col=2, text=data.get('activityToday', ''), max_rows=10, width=110)
    for row_idx in range(12, 22):
        ws.cell(row=row_idx, column=2).alignment = report_style
        ws.cell(row=row_idx, column=2).font = normal_font

    # Work Plan Next Day (G12 to G21)
    write_wrapped_rows(ws, start_row=12, col=7, text=data.get('workPlanNextDay', ''), max_rows=10, width=110)
    for row_idx in range(12, 22):
        ws.cell(row=row_idx, column=7).alignment = report_style
        ws.cell(row=row_idx, column=7).font = normal_font

def fill_team_tables(ws, data):
    mgmt_team = data.get('managementTeam', [])
    work_team = data.get('workingTeam', [])
    
    start_row = 25
    base_rows = 6
    needed_rows = max(base_rows, len(mgmt_team), len(work_team))
    shift = max(0, needed_rows - base_rows)

    # --- STEP 1: CLEAR EXISTING MERGES IN THE DANGER ZONE ---
    # This removes any pre-existing merges from the template in the table area
    # so they don't interfere with our new rows.
    target_range = f"B{start_row}:K{start_row + needed_rows}"
    
    # We have to iterate backwards through the merges to safely remove them
    for merge in list(ws.merged_cells.ranges):
        if merge.coord.startswith('B') or merge.coord.startswith('G'):
            # If the merge overlaps our table area, kill it.
            if merge.min_row >= start_row:
                ws.unmerge_cells(str(merge))
    
    # 1. Insert rows if we have more than 6 members
    if shift > 0:
        # Insert at the end of the base table (Row 31)
        ws.insert_rows(start_row + base_rows, amount=shift)
        
    # 2. STYLE & DATA LOOP (Moved OUTSIDE the shift block)
    for i in range(needed_rows):
        r = start_row + i
        ws.row_dimensions[r].height = ws.row_dimensions[15].height

        mgmt = mgmt_team[i] if i < len(mgmt_team) else {}
        work = work_team[i] if i < len(work_team) else {}

        # Ensure Merges for every data row
        try:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3) # B:C
            ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8) # G:H
        except:
            pass 

        # Style Cloning (Always copy from Row 25 to ensure consistency)
        for col in range(2, 12):
            source_cell = ws.cell(row=25, column=col)
            target_cell = ws.cell(row=r, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.alignment = copy(source_cell.alignment)

        # Fill Data
        ws.cell(row=r, column=2).value = mgmt.get('description', '') 
        ws.cell(row=r, column=4).value = mgmt.get('prev', 0)         
        ws.cell(row=r, column=5).value = mgmt.get('today', 0)        
        ws.cell(row=r, column=6).value = f"=D{r}+E{r}"

        ws.cell(row=r, column=7).value = work.get('description', '') 
        ws.cell(row=r, column=9).value = work.get('prev', 0)         
        ws.cell(row=r, column=10).value = work.get('today', 0)       
        ws.cell(row=r, column=11).value = f"=I{r}+J{r}"

    # --- NEW: APPLY DASH FORMATTING TO DATA ROWS ---
        centered_dash_format = '#,##0;(#,##0);"-"'
        for col_idx in [4, 5, 6, 9, 10, 11]:
            cell = ws.cell(row=r, column=col_idx)
            cell.number_format = centered_dash_format
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 3. TOTAL ROW LOGIC
    total_row_idx = start_row + needed_rows
    
    # Apply Merges to Total Row
    try:
        ws.merge_cells(start_row=total_row_idx, start_column=2, end_row=total_row_idx, end_column=3)
        ws.merge_cells(start_row=total_row_idx, start_column=7, end_row=total_row_idx, end_column=8)
    except:
        pass

    ws.cell(row=total_row_idx, column=2).value = "TOTAL"
    ws.cell(row=total_row_idx, column=7).value = "TOTAL"
    
    # Final Formulas
    ws.cell(row=total_row_idx, column=4).value = f"=SUM(D{start_row}:D{total_row_idx-1})"
    ws.cell(row=total_row_idx, column=5).value = f"=SUM(E{start_row}:E{total_row_idx-1})"
    ws.cell(row=total_row_idx, column=6).value = f"=SUM(F{start_row}:F{total_row_idx-1})"
    ws.cell(row=total_row_idx, column=9).value = f"=SUM(I{start_row}:I{total_row_idx-1})"
    ws.cell(row=total_row_idx, column=10).value = f"=SUM(J{start_row}:J{total_row_idx-1})"
    ws.cell(row=total_row_idx, column=11).value = f"=SUM(K{start_row}:K{total_row_idx-1})"

    # --- NEW: APPLY DASH FORMATTING TO TOTAL ROW ---
    for col_idx in [4, 5, 6, 9, 10, 11]:
        cell = ws.cell(row=r, column=col_idx)
        cell.number_format = centered_dash_format
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Final Style Fix for Total Row (Copy from the original template total row)
    original_total_row = 31 # This is where 'TOTAL' was in the empty template
    for col in range(2, 12):
        source_total = ws.cell(row=original_total_row, column=col)
        target_total = ws.cell(row=total_row_idx, column=col)
        if source_total.has_style:
            target_total.font = copy(source_total.font)
            target_total.border = copy(source_total.border)
            target_total.fill = copy(source_total.fill)
            target_total.alignment = copy(source_total.alignment)

    # 2. NOW APPLY BOLD (The "Final Layer")
    for col in range(2, 12):
        cell = ws.cell(row=total_row_idx, column=col)
        # Re-assign the font with bold=True to ensure it sticks
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True, color=cell.font.color)
    
    return shift

def fill_material_machinery_tables(ws, data, shifted_offset=0):
    materials = data.get('materials', [])
    machinery = data.get('machinery', [])
    
    # 1. POSITIONS
    base_header_row = 32
    base_sub_header_row = 33
    base_data_start = 34
    
    header_row = base_header_row + shifted_offset
    sub_header_row = base_sub_header_row + shifted_offset
    current_data_start = base_data_start + shifted_offset
    
    base_template_rows = 6 
    needed_rows = max(len(materials), len(machinery), base_template_rows)

    # 2. INSERT ROWS FIRST
    if needed_rows > base_template_rows:
        rows_to_insert = needed_rows - base_template_rows
        # FIXED: Changed current_start to current_data_start
        ws.insert_rows(current_data_start + base_template_rows, amount=rows_to_insert)

    # 3. FIX THE MAIN HEADER
    ws.row_dimensions[header_row].height = 20
    
    # Clear ghost merges
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == header_row or merge.min_row == sub_header_row: # Added sub_header cleanup
            ws.unmerge_cells(str(merge))

    # 3.2 APPLY HEADER MERGES & TEXT
    ws.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=6)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=11)
    
    # 3.3 STYLE THE MAIN HEADER (Materials / Machinery)
    # Using a professional dark blue hex: 4472C4
    header_fill = PatternFill(start_color="657C9C", end_color="657C9C", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF") # White

    for col in range(2, 12):
        target_cell = ws.cell(row=header_row, column=col)
        source_header = ws.cell(row=32, column=col) # The original template Row 32
        copy_cell_style(source_header, target_cell)
        
        # 2. Force the Colors
        target_cell.fill = header_fill
        target_cell.font = header_font
        target_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Assign text AFTER styling to ensure it shows up
    ws.cell(row=header_row, column=2).value = "Materials Deliveries"
    ws.cell(row=header_row, column=7).value = "Machinery & Equipment"

    # 4. FIX SUB-HEADER
    # This defines the text for all 10 columns (B through K)
    labels = ["Description", "Unit", "Up to Previous", "Today", "Accumulated", 
              "Description", "Unit", "Up to Previous", "Today", "Accumulated"]
    
    for i, col in enumerate(range(2, 12)):
        target_sub = ws.cell(row=sub_header_row, column=col)
        source_sub = ws.cell(row=33, column=col)
        
        # Reuse your function
        copy_cell_style(source_sub, target_sub)
        
        # Now labels[i] will work!
        target_sub.value = labels[i]
        target_sub.font = Font(name=target_sub.font.name, size=10, color="000000")
        target_sub.alignment = Alignment(horizontal='center', vertical='center')

    # 5. DATA FILLING
    for i in range(needed_rows):
        r = current_data_start + i
        ws.row_dimensions[r].height = 18
        
        mat = materials[i] if i < len(materials) else {}
        mach = machinery[i] if i < len(machinery) else {}

        for col in range(2, 12):
            copy_cell_style(ws.cell(row=34, column=col), ws.cell(row=r, column=col))

        # 1. Define the format that allows centering
        centered_dash_format = '#,##0;(#,##0);"-"'

        # --- MATERIALS (Columns D, E, F) ---
        if mat.get('description'):
            ws.cell(row=r, column=2).value = mat.get('description')
            ws.cell(row=r, column=3).value = mat.get('unit')
            ws.cell(row=r, column=4).value = to_num(mat.get('prev'))
            ws.cell(row=r, column=5).value = to_num(mat.get('today'))
            ws.cell(row=r, column=6).value = f"=D{r}+E{r}"
            
            # UPDATED: Added Column 3 to the alignment loop
            for col_idx in [3, 4, 5, 6]: 
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Only apply the number format to the numeric columns (4,5,6)
                if col_idx > 3:
                    cell.number_format = centered_dash_format

        # --- MACHINERY (Columns I, J, K) ---
        if mach.get('description'):
            ws.cell(row=r, column=7).value = mach.get('description')
            ws.cell(row=r, column=8).value = mach.get('unit')
            ws.cell(row=r, column=9).value = to_num(mach.get('prev'))
            ws.cell(row=r, column=10).value = to_num(mach.get('today'))
            ws.cell(row=r, column=11).value = f"=I{r}+J{r}"

            # Apply dash formatting for columns I, J, and K
            for col_idx in [8, 9, 10, 11]:
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Only apply the number format to the numeric columns (4,5,6)
                if col_idx > 3:
                    cell.number_format = centered_dash_format

    # 6. COLOR THE DYNAMIC FLOOR
    base_template_end = 75 
    extra_growth = max(0, needed_rows - base_template_rows)
    current_final_row = base_template_end + shifted_offset + extra_growth
    
    print(f"DEBUG: Coloring footer bar at row: {current_final_row}")
    
    footer_fill = PatternFill(start_color="657C9C", end_color="657C9C", fill_type="solid")
    
    for col in range(2, 12):
        cell = ws.cell(row=current_final_row, column=col)
        cell.fill = footer_fill
        
        # FIX: We MUST copy() the border to avoid the StyleProxy error
        source_cell = ws.cell(row=base_template_end, column=col)
        if source_cell.has_style:
            cell.border = copy(source_cell.border)
            cell.alignment = copy(source_cell.alignment)

    ws.row_dimensions[current_final_row].height = 15

    # 1. Set the View Mode to Page Break Preview (so user sees the boundary)
    ws.sheet_view.view = 'pageBreakPreview'

    # 2. Define the Print Area (From B1 to K and your dynamic footer)
    # We start at B because you want to leave Column A alone
    ws.print_area = f'B1:K{current_final_row}'

    # 3. Force "Fit to 1 Page" Logic
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 4. Center the report on the physical page
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = False # Top-aligned looks better for reports

    # 5. Set Paper Size (e.g., A4 or Letter)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Return the new total offset
    return shifted_offset + extra_growth

def fill_report_sheet(ws, data):
    """Fill the main report sheet with form data"""
    
    # Fill project info
    ws['B7'] = f"Project Name : {data.get('projectName', '')}"
    ws['B8'] = f"Weather          : AM {data.get('weatherAM', '')}  |  PM {data.get('weatherPM', '')}"
    ws['B9'] = f"Temperature  : AM {data.get('tempAM', '')}°C    |  PM {data.get('tempPM', '')}°C"
    
    # Fill activities
    ws['B12'] = data.get('activityToday', '')
    ws['G12'] = data.get('workPlanNextDay', '')
    
    # Fill teams
    fill_team_tables(ws, {
        'managementTeam': data.get('managementTeam', []),
        'workingTeam': data.get('workingTeam', [])
    })
    
    # Fill materials/machinery
    fill_material_machinery_tables(ws, {
        'materials': data.get('materials', []),
        'machinery': data.get('machinery', [])
    })