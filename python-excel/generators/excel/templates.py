from copy import copy
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell 

def copy_row(ws, src_row, tgt_row):
    ws.row_dimensions[tgt_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        tgt_cell = ws.cell(row=tgt_row, column=col)

        # THIS IS THE FIX: Prevents the "Read-only" crash
        if not isinstance(tgt_cell, MergedCell):
            tgt_cell.value = src_cell.value
        
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.protection = copy(src_cell.protection)


def copy_merged_cells(ws, src_start, src_end, offset):
    """
    Copies merged cell structures and clears ghost values 
    in the target range to ensure a clean merge.
    """
    # Use a list to avoid issues while iterating
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged in merged_ranges:
        # If the merge is within our template rows (6-9)
        if src_start <= merged.min_row <= src_end:
            new_min_row = merged.min_row + offset
            new_max_row = merged.max_row + offset
            
            # 1. CLEAR GHOST VALUES: Clear cells B and D before merging
            # This prevents the "first entry text" from appearing in the second entry
            for r in range(new_min_row, new_max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(row=r, column=c).value = None

            # 2. APPLY MERGE: Using the correct end_column parameter
            ws.merge_cells(
                start_row=new_min_row, 
                start_column=merged.min_col,
                end_row=new_max_row, 
                end_column=merged.max_col
            )

def copy_cell_style(source_cell, target_cell):
    """Copies all styling from one cell to another."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)